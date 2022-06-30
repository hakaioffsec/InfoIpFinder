# -*- coding: utf-8 -*-
import requests, json
from openpyxl import Workbook
from optparse import OptionParser
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

parser = OptionParser()
parser.add_option("-i", "--iplist", action="store", type="string", dest="iplist",
                  help="[Obrigatorio] Lista de ips", metavar="IPFILE")

parser.add_option("-x", "--xlsxfile", action="store", type="string", dest="xlsxfile",
                  help="[Obrigatorio] Manda o output pro arquivo xlsx desejado ", metavar="XLSXFILE")

parser.add_option("-c", "--collect", action="store", type="string", dest="collectmet",
                  help="[Obrigatorio] dados que deseja coletar, opcoes: a = all, b= basics, c= somente Pais de origem", metavar="Collect")

parser.add_option("-q", "--quiet",
                  action="store_false", dest="verbose", default=True,
                  help="don't print status messages to stdout")

(options, args) = parser.parse_args()




wb= Workbook()
file = open(options.iplist, "r")
cel = 1

for linha in file:

#consulta o trem
    ip = linha.strip()
    url = "http://ipinfo.io/"+ip+"?token="
    r = requests.get(url)
    resp = r.text
    print (resp)
    parse = json.loads(r.text)
#consulta outro trem
    url = "https://api11.scamalytics.com/stone.com.br/?key=&ip="+ip
    r = requests.get(url)
    resp = r.text
    parsedois = json.loads(r.text)

#paeseia as infos
    ipcap = parse['ip']
    hostname = parse['hostname']
    timezone = parse['timezone']
    orga = parse['org']
    regiao = parse['region']
    cidade = parse['city']
    localizacao = parse['loc']
    pais = parse['country']
    risco = parsedois['risk']
    score = parsedois['score']


    planilha1 = wb.active
    cel = cel + 1

#escreve no bang

    if "c" in options.collectmet:
        planilha1.cell(row=cel, column=4, value=pais)
    if "b" in options.collectmet:
        planilha1.cell(row=cel, column=4, value=pais)
        planilha1.cell(row=cel, column=1, value=ipcap)
        planilha1.cell(row=cel, column=2, value=cidade)
        planilha1.cell(row=cel, column=3, value=regiao)
        planilha1.cell(row=cel, column=5, value=localizacao)
        planilha1.cell(row=cel, column=6, value=orga)
    if "a" in options.collectmet:
        planilha1.cell(row=cel, column=4, value=pais)
        planilha1.cell(row=cel, column=1, value=ipcap)
        planilha1.cell(row=cel, column=2, value=cidade)
        planilha1.cell(row=cel, column=3, value=regiao)
        planilha1.cell(row=cel, column=5, value=localizacao)
        planilha1.cell(row=cel, column=6, value=orga)
        planilha1.cell(row=cel, column=7, value=hostname)
        planilha1.cell(row=cel, column=8, value=timezone)
        planilha1.cell(row=cel, column=9, value=risco)
        planilha1.cell(row=cel, column=10, value=score)


font = Font(name='Calibri',
                size=14,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

fill = PatternFill(fill_type='solid',
                start_color='00808080',
                end_color='00808080')

alignment=Alignment(horizontal='center',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=True,
                    indent=0)

#coloca os titulos e estiliza eles
cel = 1
if "c" in options.collectmet:
    planilha1.cell(row=cel, column=1, value="Country")
    planilha1.column_dimensions["A"].width = 20
    planilha1["A1"].font = font
    planilha1["A1"].fill = fill


if "b" in options.collectmet:
    planilha1.cell(row=cel, column=4, value="Country")
    planilha1.cell(row=cel, column=1, value='IpAddr')
    planilha1.cell(row=cel, column=2, value="Cidade")
    planilha1.cell(row=cel, column=3, value="Regiao")
    planilha1.cell(row=cel, column=5, value="Localizacao")
    planilha1.cell(row=cel, column=6, value="Org")
    planilha1.column_dimensions["A"].width = 20
    planilha1.column_dimensions["B"].width = 20
    planilha1.column_dimensions["C"].width = 20
    planilha1.column_dimensions["D"].width = 10
    planilha1.column_dimensions["E"].width = 40
    planilha1.column_dimensions["F"].width = 40
    planilha1["A1"].font = font
    planilha1["B1"].font = font
    planilha1["C1"].font = font
    planilha1["D1"].font = font
    planilha1["E1"].font = font
    planilha1["F1"].font = font
    planilha1["A1"].fill = fill
    planilha1["B1"].fill = fill
    planilha1["C1"].fill = fill
    planilha1["D1"].fill = fill
    planilha1["E1"].fill = fill
    planilha1["F1"].fill = fill




if "a" in options.collectmet:
    planilha1.cell(row=cel, column=4, value="Country")
    planilha1.cell(row=cel, column=1, value='IpAddr')
    planilha1.cell(row=cel, column=2, value="Cidade")
    planilha1.cell(row=cel, column=3, value="Regiao")
    planilha1.cell(row=cel, column=5, value="Localizacao")
    planilha1.cell(row=cel, column=6, value="Org")
    planilha1.cell(row=cel, column=7, value="Hostname")
    planilha1.cell(row=cel, column=8, value="Timezone")
    planilha1.cell(row=cel, column=9, value="Risco")
    planilha1.cell(row=cel, column=10, value="Score")
    planilha1.column_dimensions["A"].width = 20
    planilha1.column_dimensions["B"].width = 20
    planilha1.column_dimensions["C"].width = 30
    planilha1.column_dimensions["D"].width = 10
    planilha1.column_dimensions["E"].width = 40
    planilha1.column_dimensions["F"].width = 40
    planilha1.column_dimensions["G"].width = 40
    planilha1.column_dimensions["H"].width = 40
    planilha1.column_dimensions["I"].width = 40
    planilha1.column_dimensions["J"].width = 40
    planilha1["A1"].font = font
    planilha1["B1"].font = font
    planilha1["C1"].font = font
    planilha1["D1"].font = font
    planilha1["E1"].font = font
    planilha1["F1"].font = font
    planilha1["G1"].font = font
    planilha1["H1"].font = font
    planilha1["I1"].font = font
    planilha1["J1"].font = font
    planilha1["A1"].fill = fill
    planilha1["B1"].fill = fill
    planilha1["C1"].fill = fill
    planilha1["D1"].fill = fill
    planilha1["E1"].fill = fill
    planilha1["F1"].fill = fill
    planilha1["G1"].fill = fill
    planilha1["H1"].fill = fill
    planilha1["I1"].fill = fill
    planilha1["J1"].fill = fill

planilha1.alignment = alignment

wb.save(filename = options.xlsxfile)
print ("[+] Arquivo escrito com sucesso, {}".format(options.xlsxfile))

file.close()
Footer
© 2022 GitHub, Inc.
Footer navigation
Terms
Privacy
Security
Status
Docs
Contact GitHub
Pricing
API
Training
Blog
About
